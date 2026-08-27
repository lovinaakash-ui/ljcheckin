# ============================================================
# SCREENER -> EXCEL
# FAST BACKEND VERSION
# - No browser window on the happy path
# - Parallel page fetches via threads
# - Playwright only when the screener columns actually need reconfiguring
# ============================================================

import os
import pickle
import re
import sys
import time
import concurrent.futures
from pathlib import Path

import requests
from bs4 import BeautifulSoup

try:
    from playwright.sync_api import sync_playwright
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
except ImportError:
    sync_playwright = None
    PlaywrightTimeoutError = Exception

try:
    import xlwings as xw
except ImportError:
    xw = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


# ============================================================
# CONFIGURATION
# ============================================================

EMAIL = os.environ.get("SCREENER_EMAIL", "lovin.aakash@gmail.com")
PASSWORD = os.environ.get("SCREENER_PASSWORD", "Moretime01")

SCREEN_1 = "https://www.screener.in/screens/3911064/kabeer/"
SCREEN_2 = "https://www.screener.in/screens/3911102/kabeer/"

OUTPUT_FILE = Path.cwd() / "results.xlsx"
SESSION_FILE = Path.cwd() / "screener_session.pkl"
COLUMNS_MARKER = Path.cwd() / "columns_configured.pkl"

TIMEOUT = 30

# ponytail: 6 is a safe default; bump if your network is fast and screens are big
MAX_WORKERS = 6

DESIRED_COLUMNS = [
    "Return over 1day",
]

TOP_N = 3  # only keep top 3 rows per screen

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
}

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "8880948877:AAFJ1NFS4pwiqB8gfYeB2VVaDJlL2R_29Mk")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "495864674")


# ============================================================
# SESSION
# ============================================================

def create_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def save_session(session):
    with open(SESSION_FILE, "wb") as f:
        pickle.dump(requests.utils.dict_from_cookiejar(session.cookies), f)
    print("Login session saved.")


def load_session():
    if not SESSION_FILE.exists():
        return None
    try:
        s = create_session()
        with open(SESSION_FILE, "rb") as f:
            s.cookies = requests.utils.cookiejar_from_dict(pickle.load(f))
        print("Saved login session loaded.")
        return s
    except Exception as e:
        print("Could not load saved session:", e)
        return None


def login():
    print("\nLogging into Screener...")
    s = create_session()
    r = s.get("https://www.screener.in/login/", timeout=TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    csrf = (soup.find("input", {"name": "csrfmiddlewaretoken"}) or {}).get("value", "")
    r = s.post(
        "https://www.screener.in/login/",
        data={"username": EMAIL, "password": PASSWORD, "csrfmiddlewaretoken": csrf},
        headers={"Referer": "https://www.screener.in/login/"},
        allow_redirects=True,
        timeout=TIMEOUT,
    )
    r.raise_for_status()
    if "/login/" in r.url:
        test = s.get(SCREEN_1, timeout=TIMEOUT)
        if "/login/" in test.url:
            raise RuntimeError("Screener login failed.")
    print("Login successful.")
    save_session(s)
    return s


def session_is_valid(session):
    try:
        r = session.get(SCREEN_1, timeout=TIMEOUT, allow_redirects=True)
        if "/login/" in r.url:
            return False
        return r.status_code == 200 and ("Search Query" in r.text or "Run this Query" in r.text)
    except Exception:
        return False


def get_session():
    s = load_session()
    if s and session_is_valid(s):
        print("Saved session is still valid.")
        return s
    if s:
        print("Saved session expired.")
    return login()


# ============================================================
# TELEGRAM
# ============================================================

def send_telegram(message):
    """Send a message via Telegram bot API."""
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"},
            timeout=10,
        )
        if r.status_code == 200:
            print("  Telegram message sent.")
        else:
            print(f"  Telegram failed: {r.status_code} {r.text[:200]}")
    except Exception as e:
        print(f"  Telegram error: {e}")


def send_data_to_telegram(headers, all_rows):
    """Send columns C, D, E (indices 2,3,4) of each row to Telegram."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    lines = [f"<b>📊 Screener Update — {timestamp}</b>\n"]

    for i, row in enumerate(all_rows, 1):
        # Columns C, D, E = indices 2, 3, 4 (after Timestamp col A, header col B)
        c = row[1] if len(row) > 1 else "N/A"
        d = row[2] if len(row) > 2 else "N/A"
        e = row[3] if len(row) > 3 else "N/A"
        lines.append(f"<b>#{i}</b> | {c} | {d} | {e}")

    send_telegram("\n".join(lines))


# ============================================================
# FAST FETCH (threaded; I/O bound, not CPU bound)
# ============================================================

def _fetch_one(session, url):
    r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
    r.raise_for_status()
    if "/login/" in r.url:
        raise RuntimeError("SESSION_EXPIRED")
    return r.text


def _columns_present(html):
    """True if every desired column already shows in the table head."""
    soup = BeautifulSoup(html, "html.parser")
    thead = soup.find("thead")
    if not thead:
        return False
    have = [c.get_text(" ", strip=True) for c in thead.find_all(["th", "td"])]
    return all(d in have for d in DESIRED_COLUMNS)


def _total_pages_from_html(html):
    """Best-effort page count from the pagination block of page 1."""
    soup = BeautifulSoup(html, "html.parser")
    pagination = soup.select_one(".pagination")
    if not pagination:
        return 1
    max_page = 1
    for link in pagination.find_all("a"):
        m = re.search(r"page=(\d+)", link.get("href", ""))
        if m:
            max_page = max(max_page, int(m.group(1)))
        t = link.get_text(" ", strip=True)
        if t.isdigit():
            max_page = max(max_page, int(t))
    return max(max_page, 1)


def _page_url(screen_url, page):
    if page == 1:
        return screen_url
    sep = "&" if "?" in screen_url else "?"
    return f"{screen_url}{sep}page={page}"


def fetch_screen(session, screen_url, screen_number):
    """Fetch all pages of a screen; page 1 sequential, the rest in a thread pool."""
    print(f"\n{'='*60}\nSCREEN {screen_number}\n{'='*60}")

    first_html = _fetch_one(session, screen_url)
    columns_ok = _columns_present(first_html)
    total_pages = _total_pages_from_html(first_html)

    if total_pages == 1:
        htmls = [first_html]
    else:
        page_urls = [_page_url(screen_url, p) for p in range(2, total_pages + 1)]
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            htmls = [first_html] + list(ex.map(lambda u: _fetch_one(session, u), page_urls))

    headers = None
    all_rows = []
    for html in htmls:
        h, rows = parse_results(html, screen_number)
        headers = headers or h
        all_rows.extend(rows)

    print(f"Screen {screen_number} TOTAL: {len(all_rows)} companies")
    return headers, all_rows, columns_ok


# ============================================================
# PARSE
# ============================================================

def parse_results(html, screen_number):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", id="resultsTable")

    if table is None:
        # Fallback: biggest table on the page
        best = None
        best_n = 0
        for t in soup.find_all("table"):
            n = len(t.find_all("tr"))
            if n > best_n:
                best, best_n = t, n
        table = best

    if table is None:
        raise RuntimeError(f"Screen {screen_number}: Result table not found.")

    head = table.find("thead")
    if head:
        header_cells = head.find_all(["th", "td"])
    else:
        first_row = table.find("tr")
        header_cells = first_row.find_all(["th", "td"]) if first_row else []
    headers = [c.get_text(" ", strip=True) for c in header_cells]

    tbody = table.find("tbody")
    rows_html = tbody.find_all("tr") if tbody else table.find_all("tr")[1:]

    rows = []
    for tr in rows_html:
        cells = tr.find_all("td")
        if not cells:
            continue
        rows.append([c.get_text(" ", strip=True) for c in cells])

    print(f"Screen {screen_number}: {len(rows)} companies found.")
    return headers, rows


# ============================================================
# PLAYWRIGHT (only when columns are not already correct)
# ============================================================

def _browser_cookies(session):
    return [{
        "name": c.name, "value": c.value,
        "domain": c.domain or ".screener.in",
        "path": c.path or "/",
    } for c in session.cookies]


# Chips are <li class="draggable-row"> elements inside the Manage Columns dialog.
# Each contains the column name as text (no × in innerText).

CHIP_DETECT_JS = r"""
() => {
    const selected = [];
    for (const ch of document.querySelectorAll('li.draggable-row[data-name]')) {
        const name = ch.getAttribute('data-name').trim();
        if (name) selected.push(name);
    }
    return selected;
}
"""

# Each chip: <li class="draggable-row" data-name="X">...<button class="button-plain"><i class="icon-cancel-thin"/></button></li>
# The close button is: button.button-plain inside the li.
CHIP_CLICK_JS = r"""
(name) => {
    for (const ch of document.querySelectorAll('li.draggable-row[data-name]')) {
        if (ch.getAttribute('data-name').trim().toLowerCase() !== name.toLowerCase()) continue;
        const btn = ch.querySelector('button.button-plain');
        if (btn) { btn.click(); return true; }
        ch.click();
        return true;
    }
    return false;
}
"""


def _get_selected_chips(page):
    """Return list of visible selected-column chip names in the Manage Columns dialog."""
    return page.evaluate(CHIP_DETECT_JS) or []


def _click_chip_close_by_name(page, column_name):
    """Click the close button on the chip whose data-name matches column_name."""
    chip = page.locator(f'li.draggable-row[data-name="{column_name}"] button.button-plain').first
    if chip.count() > 0 and chip.is_visible():
        chip.click()
        return True
    # Fallback: try via JS
    return bool(page.evaluate(CHIP_CLICK_JS, column_name))


def _click_column_result(page, column_name):
    # Use the known search input directly — avoids looping all inputs.
    search = page.locator("input[placeholder*='return on capital' i]").first
    if not search.is_visible():
        # Broader fallback: first visible text/search input
        for i in range(page.locator("input").count()):
            c = page.locator("input").nth(i)
            try:
                if c.is_visible() and (c.get_attribute("type") or "text") in ("text", "search"):
                    search = c
                    break
            except Exception:
                pass

    search.fill("")
    search.fill(column_name)

    # Wait for the dropdown result to appear instead of a fixed timeout.
    result = page.locator(f"label.ratio").filter(has_text=column_name).first
    try:
        result.wait_for(state="visible", timeout=3000)
    except Exception:
        raise RuntimeError(f"Could not select Screener column: {column_name}")
    result.click()


def configure_columns_in_browser(session, screen_url, screen_number):
    if sync_playwright is None:
        raise RuntimeError(
            "Playwright is required for the Edit Columns step. "
            "Install it with: pip install playwright && playwright install chromium"
        )

    print(f"\n{'='*60}\nCONFIGURING COLUMNS - SCREEN {screen_number}\n{'='*60}")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context()
        try:
            cookies = _browser_cookies(session)
            if cookies:
                context.add_cookies(cookies)

            page = context.new_page()
            page.goto(screen_url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1000)
            if "/login/" in page.url:
                raise RuntimeError("SESSION_EXPIRED")

            page.get_by_text("Edit Columns", exact=True).click(timeout=15000)
            page.get_by_text("Manage columns", exact=True).wait_for(timeout=15000)
            # Wait for the chip list to actually render before reading it.
            try:
                page.wait_for_selector("li.draggable-row", timeout=5000)
            except Exception:
                pass
            page.wait_for_timeout(500)

            # Diff: keep DESIRED_COLUMNS already present, remove the rest.
            desired = {c.lower() for c in DESIRED_COLUMNS}
            selected = _get_selected_chips(page)
            to_remove = [c for c in selected if c.lower() not in desired]
            to_add    = [c for c in DESIRED_COLUMNS if c.lower() not in {s.lower() for s in selected}]

            print(f"  Already present: {[c for c in selected if c.lower() in desired]}")
            print(f"  Removing {len(to_remove)} extra column(s): {to_remove}")
            print(f"  Adding {len(to_add)} column(s): {to_add}")

            for col in to_remove:
                _click_chip_close_by_name(page, col)
                page.wait_for_timeout(250)

            if to_remove:
                page.wait_for_timeout(800)
                leftover = [c for c in _get_selected_chips(page) if c.lower() not in desired]
                if leftover:
                    page.screenshot(path=f"column_clear_not_finished_screen_{screen_number}.png", full_page=True)
                    raise RuntimeError(
                        f"Could not remove {leftover} on screen {screen_number}."
                    )

            print("Adding requested columns...")
            for column_name in to_add:
                print(f"  + {column_name}")
                _click_column_result(page, column_name)

            print("Saving columns...")
            save_clicked = False
            try:
                btn = page.locator("button, input[type='button'], input[type='submit'], a").filter(
                    has_text=re.compile(r"^\s*SAVE COLUMNS\s*$", re.I)
                ).first
                if btn.is_visible():
                    btn.click(timeout=5000)
                    save_clicked = True
            except Exception:
                pass
            if not save_clicked:
                try:
                    save_clicked = bool(page.evaluate(r"""
                        () => {
                            const el = [...document.querySelectorAll(
                                'button, input[type="button"], input[type="submit"], a'
                            )].find(e => {
                                const t = (e.innerText || e.value || '').replace(/\s+/g,' ').trim().toLowerCase();
                                const r = e.getBoundingClientRect();
                                return r.width>0 && r.height>0 && t === 'save columns';
                            });
                            if (el) { el.click(); return true; }
                            return false;
                        }
                    """))
                except Exception:
                    pass
            if not save_clicked:
                page.screenshot(path=f"column_save_failed_screen_{screen_number}.png", full_page=True)
                raise RuntimeError(f"SAVE COLUMNS could not be clicked on screen {screen_number}.")

            try:
                page.wait_for_load_state("domcontentloaded", timeout=5000)
            except Exception:
                pass
            page.wait_for_timeout(2500)

            # Sync cookies back to the requests session.
            for c in context.cookies():
                session.cookies.set(
                    c["name"], c["value"],
                    domain=c.get("domain"), path=c.get("path", "/"),
                )
            save_session(session)
        except PlaywrightTimeoutError as error:
            raise RuntimeError(
                f"Timed out while configuring columns on screen {screen_number}."
            ) from error
        finally:
            browser.close()


# ============================================================
# EXCEL OUTPUT (xlwings for live writing, openpyxl fallback)
# ============================================================

def _create_workbook(headers, rows, timestamp):
    """Create a new xlsx with openpyxl (used on first run)."""
    if Workbook is None:
        raise RuntimeError("openpyxl required: pip install openpyxl")
    for attempt in range(3):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp"] + headers)
            for row in rows:
                ws.append([timestamp] + row)
            wb.save(str(OUTPUT_FILE))
            wb.close()
            print(f"Created: {OUTPUT_FILE}")
            return
        except PermissionError:
            print(f"  File locked (attempt {attempt+1}/3), retrying in 2s...")
            time.sleep(2)
            if attempt == 2:
                raise


def _append_xlwings(headers, rows, timestamp):
    """Append rows to the open Excel workbook via xlwings COM — no file lock."""
    if xw is None:
        return False
    try:
        # Find the running Excel instance and the open workbook
        for app in xw.apps:
            for wb in app.books:
                if wb.name == OUTPUT_FILE.name:
                    ws = wb.sheets[0]
                    # Find next empty row
                    last_row = ws.range("A" + str(ws.cells.last_cell.row)).end("up").row
                    if last_row == 1 and ws.range("A1").value is None:
                        last_row = 0  # sheet is empty
                    next_row = last_row + 1
                    # Write header if sheet is empty
                    if next_row == 1:
                        for ci, h in enumerate(["Timestamp"] + headers, 1):
                            ws.cells(next_row, ci).value = h
                        next_row = 2
                    # Write data rows
                    for row in rows:
                        for ci, val in enumerate([timestamp] + row, 1):
                            ws.cells(next_row, ci).value = val
                        next_row += 1
                    wb.api.Save()
                    print(f"Appended {len(rows)} row(s) to open Excel via xlwings.")
                    return True
        return False  # workbook not found in open Excel
    except Exception as e:
        print(f"  xlwings append failed: {e}")
        return False


def _append_openpyxl_fallback(headers, rows, timestamp):
    """Fallback: read existing xlsx, add rows, write to temp, swap."""
    if load_workbook is None:
        raise RuntimeError("openpyxl required: pip install openpyxl")
    try:
        wb = load_workbook(str(OUTPUT_FILE))
        ws = wb.active
        for row in rows:
            ws.append([timestamp] + row)
        tmp = OUTPUT_FILE.with_suffix(".tmp")
        wb.save(str(tmp))
        wb.close()
        os.replace(str(tmp), str(OUTPUT_FILE))
        print(f"Appended {len(rows)} row(s) via openpyxl fallback.")
        return True
    except PermissionError:
        # Last resort: save as new file
        alt = OUTPUT_FILE.with_stem(OUTPUT_FILE.stem + "_new")
        try:
            wb = load_workbook(str(OUTPUT_FILE))
            ws = wb.active
            for row in rows:
                ws.append([timestamp] + row)
            wb.save(str(alt))
            wb.close()
            print(f"  ⚠ {OUTPUT_FILE.name} locked — saved as {alt.name}")
            return True
        except Exception:
            print(f"  Failed to save anywhere.")
            return False


def write_to_excel(headers, rows, append=False):
    """Main entry: write/append to results.xlsx. Tries xlwings first (works with Excel open)."""
    print(f"\n{'='*60}\n{'APPENDING' if append else 'CREATING'} EXCEL\n{'='*60}")
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    if not append or not OUTPUT_FILE.exists():
        _create_workbook(headers, rows, timestamp)
        return

    # Try xlwings first (works even when Excel has the file open)
    if _append_xlwings(headers, rows, timestamp):
        return

    # Fallback: openpyxl temp-swap
    _append_openpyxl_fallback(headers, rows, timestamp)


def open_excel():
    """Open the xlsx in Excel (Windows only; skipped on Linux/cloud)."""
    if sys.platform != "win32":
        return  # cloud / Linux — nothing to open
    for attempt in range(3):
        try:
            if OUTPUT_FILE.exists():
                print("Opening results.xlsx...")
                os.startfile(str(OUTPUT_FILE))
                return
            else:
                print("  File not yet created, waiting...")
                time.sleep(1)
        except PermissionError:
            print(f"  File locked (attempt {attempt+1}/3), retrying in 2s...")
            time.sleep(2)
    print("  Could not open Excel — open results.xlsx manually.")


# ============================================================
# MAIN
# ============================================================

def _run_one_screen(session, screen_url, n):
    try:
        return fetch_screen(session, screen_url, n)
    except RuntimeError as e:
        if str(e) == "SESSION_EXPIRED":
            return ("RELOGIN",)
        raise


def _columns_configured():
    if not COLUMNS_MARKER.exists():
        return False
    try:
        with open(COLUMNS_MARKER, "rb") as f:
            return pickle.load(f).get("done", False)
    except PermissionError:
        print("  Warning: columns_configured.pkl locked, treating as not configured.")
        return False
    except Exception:
        return False


def _mark_columns_configured():
    with open(COLUMNS_MARKER, "wb") as f:
        pickle.dump({"done": True}, f)


def _normalize_row(row, n_cols):
    row = list(row)
    if len(row) < n_cols:
        row = row + [""] * (n_cols - len(row))
    elif len(row) > n_cols:
        row = row[:n_cols]
    return row


def _ensure_columns(session):
    """Run Playwright column config. Skip if marker exists."""
    if _columns_configured():
        print("Columns already configured (skipping Playwright).")
        return
    print("Running Playwright column setup...")
    configure_columns_in_browser(session, SCREEN_1, 1)
    _mark_columns_configured()


def main():
    start = time.time()
    print(f"\n{'='*60}\nSCREENER FAST BACKEND EXPORT\n{'='*60}")

    session = get_session()
    _ensure_columns(session)

    # Process both screens in parallel.
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
        f1 = ex.submit(_run_one_screen, session, SCREEN_1, 1)
        f2 = ex.submit(_run_one_screen, session, SCREEN_2, 2)
        results = [f1.result(), f2.result()]

    # Re-login if any screen expired, then re-fetch that screen.
    re_needed = [i for i, r in enumerate(results) if r and r[0] == "RELOGIN"]
    if re_needed:
        print("Session expired; logging in again...")
        session = login()
        for i in re_needed:
            url, n = (SCREEN_1, 1) if i == 0 else (SCREEN_2, 2)
            results[i] = fetch_screen(session, url, n)

    h1, r1, _ = results[0]
    h2, r2, _ = results[1]

    if not r1:
        raise RuntimeError("SCREEN 1 RETURNED ZERO RESULTS.")
    if not r2:
        raise RuntimeError("SCREEN 2 RETURNED ZERO RESULTS.")

    headers = h1
    n_cols = len(headers)
    already_exists = OUTPUT_FILE.exists()

    # --- Real-time Excel: open after screen 1, update after screen 2 ---
    rows1 = [_normalize_row(r, n_cols) for r in r1[:TOP_N]]
    write_to_excel(headers, rows1, append=already_exists)
    open_excel()
    already_exists = True  # file now exists, next write should append
    print(f"\nExcel opened with Screen 1 data ({len(rows1)} rows). Waiting for Screen 2...")

    rows2 = [_normalize_row(r, n_cols) for r in r2[:TOP_N]]
    write_to_excel(headers, rows2, append=already_exists)
    print(f"Excel updated with both screens ({len(rows1) + len(rows2)} total rows).")

    # Send to Telegram
    send_data_to_telegram(headers, rows1 + rows2)

    elapsed = time.time() - start
    print(f"\n{'='*60}\nCOMPLETED\nScreen 1: {len(r1)}\nScreen 2: {len(r2)}\n"
          f"Total: {len(r1) + len(r2)}\nTime: {elapsed:.2f} seconds\n{'='*60}")


if __name__ == "__main__":
    import traceback
    once = "--once" in sys.argv
    if once:
        main()
    else:
        print("Auto-refresh every 30 seconds. Press Ctrl+C to stop.\n")
        try:
            while True:
                try:
                    main()
                except Exception as error:
                    print(f"\n{'='*60}\nERROR\n{'='*60}\n{repr(error)}")
                    traceback.print_exc()
                print(f"\nRerunning in 30 seconds... (Ctrl+C to stop)")
                time.sleep(30)
        except KeyboardInterrupt:
            print("\nStopped.")
